import tkinter as tk
from tkinter import ttk
import openpyxl
from tkinter import messagebox

def load_data():
    path = "Data.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)

        treeview.delete(*treeview.get_children())

        if list_values: 
            column_names = ["Name", "House", "Merit/Demerit Points"]
            treeview.config(columns=column_names)
            for col_name in column_names:
                treeview.heading(col_name, text=col_name)

            for value_tuple in list_values[1:]: 
                name = value_tuple[0]
                house = value_tuple[3] if value_tuple[3] else "No House"
                merit_demerit_points = value_tuple[4] if isinstance(value_tuple[4], int) else 0 
               
                formatted_points = f"+{merit_demerit_points}" if merit_demerit_points > 0 else f"{merit_demerit_points}"
                treeview.insert('', tk.END, values=(name, house, formatted_points))
        else:
            messagebox.showinfo("Info", "No data found in the Excel file.")
    except FileNotFoundError:
        messagebox.showerror("Error", f"File '{path}' not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def add_points(points_type):
    selected_item = treeview.selection()
    if not selected_item:
        messagebox.showwarning("Warning", "No student selected.")
        return

    student_name = treeview.item(selected_item)['values'][0]

    path = "Data.xlsx"
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2):  
            if row[0].value == student_name:
                current_points = row[4].value if isinstance(row[4].value, int) else 0  
                if points_type == 'merit':
                    row[4].value = current_points + 1
                    messagebox.showinfo("Success", f"Added 1 merit point to {student_name}.")
                elif points_type == 'demerit':
                    row[4].value = current_points - 1
                    messagebox.showinfo("Success", f"Added 1 demerit point to {student_name}.")
                break

        workbook.save(path)
        load_data()  

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating points: {e}")

def filter_students_by_class(event):
    selected_class = class_combobox.get()
    if selected_class:
        for item in treeview.get_children():
            treeview.delete(item)  

        path = "Data.xlsx"
        try:
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            list_values = list(sheet.values)

            if list_values:
                for value_tuple in list_values[1:]:
                    if value_tuple[1] == selected_class:  
                        name = value_tuple[0]
                        house = value_tuple[3] if value_tuple[3] else "No House"
                        merit_demerit_points = value_tuple[4] if isinstance(value_tuple[4], int) else 0 
                        
                        formatted_points = f"+{merit_demerit_points}" if merit_demerit_points > 0 else f"{merit_demerit_points}"
                        treeview.insert('', tk.END, values=(name, house, formatted_points))
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while filtering: {e}")

def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

root = tk.Tk()
root.title("Merit and Demerit Points System")

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

class_list = ["5 - Blue", "5 - Green", "6 - Blue", "6 - Green", "7 - Blue", "7 - Green", "8 - Blue", "8 - Green", "9 - Blue", "9 - Green", "10 - Blue", "10 - Green", "11 - Blue", "11 - Green"]

frame = ttk.Frame(root)
frame.pack()

left_frame = ttk.Frame(frame)
left_frame.pack(side=tk.LEFT, padx=20, pady=10)

title_label = ttk.Label(left_frame, text="Merit/Demerit Points", font=("Arial", 14, "bold"))
title_label.grid(row=0, column=0, sticky="w")

class_combobox = ttk.Combobox(left_frame, values=class_list)
class_combobox.current(0)
class_combobox.bind("<<ComboboxSelected>>", filter_students_by_class)
class_combobox.grid(row=1, column=0, padx=5, pady=(5, 10), sticky="ew")

treeFrame = ttk.Frame(frame, width=200)  # Set a specific width for the treeFrame
treeFrame.pack(side=tk.RIGHT, pady=10, fill=tk.Y)  # Prevent horizontal expansion

treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Name", "House", "Merit/Demerit Points")
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=13)

# Adjust the width of each column
treeview.column("Name", width=80)  # Adjust width as needed
treeview.column("House", width=80)
treeview.column("Merit/Demerit Points", width=100)

for col in cols:
    treeview.heading(col, text=col)

treeview.pack(fill=tk.BOTH)
treeScroll.config(command=treeview.yview)


load_data()

mode_switch = ttk.Checkbutton(
    left_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=2, column=0, padx=5, pady=10, sticky="nsew")

add_merit_button = ttk.Button(left_frame, text="Add Merit Point", command=lambda: add_points('merit'))
add_merit_button.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

add_demerit_button = ttk.Button(left_frame, text="Add Demerit Point", command=lambda: add_points('demerit'))
add_demerit_button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

root.mainloop()
