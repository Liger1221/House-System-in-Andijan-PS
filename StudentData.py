import tkinter as tk
from tkinter import ttk
import openpyxl
from tkinter import messagebox

def load_data():
    path = "Data.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)

        # Clear the Treeview
        treeview.delete(*treeview.get_children())

        if list_values: 
            # Define the columns we want to display
            column_names = ["Name", "Class", "Average Score", "House"]
            for col_name in column_names:
                treeview.heading(col_name, text=col_name)
                treeview.column(col_name, width=100)

            # Insert data only for the specified columns
            for value_tuple in list_values[1:]:
                # Only add the desired values
                if all(value_tuple[i] is not None for i in [0, 1, 2]):  # Ensure no None values for Name, Class, Average Score
                    filtered_values = (value_tuple[0], value_tuple[1], value_tuple[2], value_tuple[3])  # Adjust this based on your actual column indices
                    treeview.insert('', tk.END, values=filtered_values)
        else:
            messagebox.showinfo("Info", "No data found in the Excel file.")
    except FileNotFoundError:
        messagebox.showerror("Error", f"File '{path}' not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
def insert_row():
    name = name_entry.get()
    student_class = class_combobox.get()

    try:
        average_score = float(score_entry.get())
        if average_score > 100:
            messagebox.showerror("Input Error", "Average Score cannot exceed 100. Please enter a valid number.")
            return
    except ValueError:
        messagebox.showerror("Input Error", "Invalid input for Average Score. Please enter a valid number.")
        return

    house = "None"  # Initially set house as 'None'

    path = "Data.xlsx"
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [name, student_class, average_score, house]
        sheet.append(row_values)
        workbook.save(path)

        # Insert the new row into the Treeview
        treeview.insert('', tk.END, values=row_values)

        # Reset the input fields
        name_entry.delete(0, "end")
        name_entry.insert(0, "Name")
        class_combobox.set(class_list[0])
        score_entry.delete(0, "end")
        score_entry.insert(0, "Average Score")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving to Excel: {e}")
def delete_selected_student():
    selected_item = treeview.selection()
    if not selected_item:
        messagebox.showwarning("Warning", "No student selected to delete.")
        return

    
    student_name = treeview.item(selected_item)['values'][0]
    student_class = treeview.item(selected_item)['values'][1]

    
    for item in selected_item:
        treeview.delete(item)

    
    path = "Data.xlsx"
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
       
        for row in sheet.iter_rows(min_row=2): 
            if row[0].value == student_name and row[1].value == student_class:
                sheet.delete_rows(row[0].row)  
                break
        
        workbook.save(path)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while deleting from Excel: {e}")

def filter_students_by_class(event):
    selected_class = class_combobox.get()
    if selected_class:
        for item in treeview.get_children():
            treeview.delete(item) 

        path = "Data.xlsx"
        try:
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            list_values = list(sheet.values)

            if list_values:
                for value_tuple in list_values[1:]:
                    if value_tuple[1] == selected_class:  # Check the class
                        treeview.insert('', tk.END, values=value_tuple)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while filtering: {e}")
def distribute_houses():
    path = "Data.xlsx"
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        list_values = list(sheet.values)

        selected_class = class_combobox.get() 

        if len(list_values) < 2:
            messagebox.showinfo("Info", "Not enough students to distribute.")
            return

        students = [] 

        
        for value_tuple in list_values[1:]:
            if value_tuple[1] == selected_class: 
                students.append((value_tuple[0], value_tuple[1], value_tuple[2])) 

        if not students:
            messagebox.showinfo("Info", f"No students found in {selected_class}.")
            return

        houses = ["Blue House", "Red House", "Green House", "White House"]

        students.sort(key=lambda x: x[2], reverse=True)

        for i, student in enumerate(students):
            student_house = houses[i % 4] 

            student_with_house = (student[0], student[1], student[2], student_house)

            for row in sheet.iter_rows(min_row=2):
                if row[0].value == student_with_house[0] and row[1].value == student_with_house[1]: 
                    sheet.cell(row=row[0].row, column=4, value=student_house)

        workbook.save(path)
        load_data() 
        messagebox.showinfo("Success", f"Students in {selected_class} have been distributed to houses.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while distributing houses: {e}")

def set_default_text(entry, default_text):
    """Sets the default text if the entry is empty."""
    if entry.get() == "":
        entry.insert(0, default_text)

def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

root = tk.Tk()
root.title("Student House Distributor")

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

class_list = ["5 - Blue", "5 - Green", "6 - Blue", "6 - Green", "7 - Blue", "7 - Green", "8 - Blue", "8 - Green", "9 - Blue", "9 - Green", "10 - Blue", "10 - Green", "11 - Blue", "11 - Green"]

frame = ttk.Frame(root)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

class_combobox = ttk.Combobox(widgets_frame, values=class_list)
class_combobox.current(0)
class_combobox.bind("<<ComboboxSelected>>", filter_students_by_class)
class_combobox.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Name")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.bind("<FocusOut>", lambda e: set_default_text(name_entry, "Name"))
name_entry.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

score_entry = ttk.Entry(widgets_frame)
score_entry.insert(0, "Average Score")
score_entry.bind("<FocusIn>", lambda e: score_entry.delete('0', 'end'))
score_entry.bind("<FocusOut>", lambda e: set_default_text(score_entry, "Average Score"))
score_entry.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="ew")

button = ttk.Button(widgets_frame, text="Insert", command=insert_row)
button.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

sort_button = ttk.Button(widgets_frame, text="Distribute Houses", command=distribute_houses)
sort_button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

delete_button = ttk.Button(widgets_frame, text="Delete Selected", command=delete_selected_student)
delete_button.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=6, column=0, padx=(20, 10), pady=10, sticky="ew")

mode_switch = ttk.Checkbutton(
    widgets_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=7, column=0, padx=5, pady=10, sticky="nsew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Name", "Class", "Average Score", "House")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=13)

for col in cols:
    treeview.heading(col, text=col)
    treeview.column(col, width=100)

treeview.pack()
treeScroll.config(command=treeview.yview)

load_data()

root.mainloop()
