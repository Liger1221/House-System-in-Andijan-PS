import tkinter as tk
from tkinter import ttk
import openpyxl
from tkinter import messagebox

def load_points():
    path = "Houses.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        # Load points from specific cells
        blue_points = sheet['A2'].value or 0
        red_points = sheet['B2'].value or 0
        green_points = sheet['C2'].value or 0
        white_points = sheet['D2'].value or 0
        
        blue_entry.delete(0, tk.END)
        blue_entry.insert(0, blue_points)

        red_entry.delete(0, tk.END)
        red_entry.insert(0, red_points)

        green_entry.delete(0, tk.END)
        green_entry.insert(0, green_points)

        white_entry.delete(0, tk.END)
        white_entry.insert(0, white_points)

    except FileNotFoundError:
        messagebox.showerror("Error", f"File '{path}' not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def update_points(house, change):
    path = "Houses.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        if house == "Blue":
            current_points = sheet['A2'].value or 0
            sheet['A2'] = current_points + change
        elif house == "Red":
            current_points = sheet['B2'].value or 0
            sheet['B2'] = current_points + change
        elif house == "Green":
            current_points = sheet['C2'].value or 0
            sheet['C2'] = current_points + change
        elif house == "White":
            current_points = sheet['D2'].value or 0
            sheet['D2'] = current_points + change
        
        workbook.save(path)
        load_points()  # Reload the points after updating

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating points: {e}")

def save_points():
    path = "Houses.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        # Save the points from the entry fields to the Excel sheet
        sheet['A2'] = int(blue_entry.get() or 0)
        sheet['B2'] = int(red_entry.get() or 0)
        sheet['C2'] = int(green_entry.get() or 0)
        sheet['D2'] = int(white_entry.get() or 0)
        
        workbook.save(path)
        messagebox.showinfo("Success", "Changes saved successfully!")
        
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving points: {e}")

def add_points(house):
    update_points(house, 1)

def remove_points(house):
    update_points(house, -1)

def toggle_mode():
    # Toggle between light and dark themes
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
        root.configure(bg="#FFFFFF")  # Change root background to light
        
        mode_switch.configure(style="TCheckbutton")
    else:
        style.theme_use("forest-dark")
        root.configure(bg="#2E2E2E")  # Change root background to dark
    
        mode_switch.configure(style="Dark.TCheckbutton")


root = tk.Tk()
root.title("House Points Management")

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

frame = ttk.Frame(root)
frame.pack(padx=20, pady=20)

# Blue House
blue_frame = ttk.Frame(frame)
blue_frame.pack(padx=10, pady=5, fill="x")
ttk.Label(blue_frame, text="Blue:").pack(side="left", padx=(0, 20))  # Added extra space (20px)
ttk.Button(blue_frame, text="-", command=lambda: remove_points("Blue")).pack(side="left")
blue_entry = ttk.Entry(blue_frame, width=10)
blue_entry.pack(side="left", padx=5)
ttk.Button(blue_frame, text="+", command=lambda: add_points("Blue")).pack(side="left")

# Red House
red_frame = ttk.Frame(frame)
red_frame.pack(padx=10, pady=5, fill="x")
ttk.Label(red_frame, text="Red:").pack(side="left", padx=(0, 22))  # Added extra space (20px)
ttk.Button(red_frame, text="-", command=lambda: remove_points("Red")).pack(side="left")
red_entry = ttk.Entry(red_frame, width=10)
red_entry.pack(side="left", padx=5)
ttk.Button(red_frame, text="+", command=lambda: add_points("Red")).pack(side="left")

# Green House
green_frame = ttk.Frame(frame)
green_frame.pack(padx=10, pady=5, fill="x")
ttk.Label(green_frame, text="Green:").pack(side="left", padx=(0, 10))  # Added extra space (20px)
ttk.Button(green_frame, text="-", command=lambda: remove_points("Green")).pack(side="left")
green_entry = ttk.Entry(green_frame, width=10)
green_entry.pack(side="left", padx=5)
ttk.Button(green_frame, text="+", command=lambda: add_points("Green")).pack(side="left")

# White House
white_frame = ttk.Frame(frame)
white_frame.pack(padx=10, pady=5, fill="x")
ttk.Label(white_frame, text="White:").pack(side="left", padx=(0, 10))  # Added extra space (20px)
ttk.Button(white_frame, text="-", command=lambda: remove_points("White")).pack(side="left")
white_entry = ttk.Entry(white_frame, width=10)
white_entry.pack(side="left", padx=5)
ttk.Button(white_frame, text="+", command=lambda: add_points("White")).pack(side="left")

separator = ttk.Separator(frame)
separator.pack(fill="x", pady=10)

# Save button
save_button = ttk.Button(frame, text="Save Changes", command=save_points)
save_button.pack(pady=10)

mode_switch = ttk.Checkbutton(frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.pack(pady=10)

load_points()

root.mainloop()
